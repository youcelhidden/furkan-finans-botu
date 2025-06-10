import requests
import time
import schedule
from bs4 import BeautifulSoup
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
import smtplib
from email.mime.text import MIMEText

# === Telegram Ayarları ===
BOT_TOKEN = "7878308274:AAG7bxP7gfkbujj0A5L8GGKkCFVaYu_jUEw"
CHAT_ID = "1222943089"

# === Mail Ayarları ===
MAIL_GONDEREN = "yucelfurkan123@gmail.com"
MAIL_SIFRE = "xdaj mxly rfht vzfz"
MAIL_ALICI = "yucelfurkan123@gmail.com"

# === Excel Kayıt ===
def log_to_excel(data):
    filename = "piyasa_verileri.xlsx"
    headers = ["Tarih", "Euro", "Dolar", "Altın", "BIST100", "Bitcoin", "Gümüş", "Brent"]

    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.append([now] + list(data))
    wb.save(filename)

# === Mail Gönderimi ===
def send_email(message):
    msg = MIMEText(message)
    msg["Subject"] = "📬 Günlük Piyasa Özeti"
    msg["From"] = MAIL_GONDEREN
    msg["To"] = MAIL_ALICI

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(MAIL_GONDEREN, MAIL_SIFRE)
            server.sendmail(MAIL_GONDEREN, MAIL_ALICI, msg.as_string())
    except Exception as e:
        print("Mail gönderimi başarısız:", e)

# === Telegram Mesaj Gönderimi ===
def send_telegram_message(message):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": CHAT_ID,
        "text": message
    }
    requests.post(url, data=payload)

# === Veri Çekme ===
def get_market_data():
    url = "https://www.doviz.com/"
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")

    def get_text(key, attr="s"):
        tag = soup.select_one(f'span[data-socket-key="{key}"][data-socket-attr="{attr}"]')
        if tag:
            value = tag.text.strip().replace(".", "").replace(",", ".")
            return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return "YOK"

    try:
        euro    = get_text("EUR")
        dolar   = get_text("USD")
        altin   = get_text("gram-altin")
        borsa   = get_text("XU100")
        bitcoin = get_text("bitcoin")
        gumus   = get_text("gumus")
        brent   = get_text("brent")
        return euro, dolar, altin, borsa, bitcoin, gumus, brent
    except:
        return ["HATA"] * 7

# === Piyasa Bülteni ===
def piyasa_bulteni():
    euro, dolar, altin, borsa, bitcoin, gumus, brent = get_market_data()
    now = datetime.now().strftime("%d.%m.%Y")

    if euro == "YOK" or dolar == "YOK" or borsa == "YOK":
        hata_mesaj = f"⚠️ Veri çekilemedi ({now}) – Lütfen bağlantıyı kontrol et."
        send_telegram_message(hata_mesaj)
        send_email(hata_mesaj)
        return

    message = f"""📬 Günlük Piyasa Özeti ({now})
💶 Euro: {euro} TL
💵 Dolar: {dolar} TL
🥇 Gram Altın: {altin} TL
📈 BIST 100: {borsa}
🪙 Bitcoin: ${bitcoin}
🥈 Gümüş: {gumus} TL
🛢️ Brent: ${brent}"""

    send_telegram_message(message)
    send_email(message)
    log_to_excel((euro, dolar, altin, borsa, bitcoin, gumus, brent))

# === Zamanlayıcı ===
schedule.every().day.at("10:00").do(piyasa_bulteni)
schedule.every().day.at("18:30").do(piyasa_bulteni)

print("📡 Genişletilmiş Piyasa Botu aktif...")

while True:
    schedule.run_pending()
    time.sleep(60)
